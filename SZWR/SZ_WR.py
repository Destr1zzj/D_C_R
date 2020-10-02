# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import os
import ast
import time

"""
read me!
class SZ_W(DMG_dict, zongpai)
    write_2_excel()  写入，需要传入伤害【dict】zongpai【int】
非class    
read_from_excel(zongpai, school, cols_plus=0) 读取单个门派数据， 需要传入 宗派int，门派int，读取列数int（默认传0）
read_from_excel_all(zongpai, cols_plus=0)读取某宗派的全部数据，需要传入 宗派int，读取列数int（默认传0）
log(*err_data)  打啥传啥就可，会自动加日期

"""
class SZ_W():

    def __init__(self,DMG_dict, zongpai):
        '''
        zongpai 参数（0，1，2）
        即对应登录后从上到下1，2，3个账号
        self.zp 宗派012
        self.data 伤害dict
        self.schoool 门派1-8(不需要传入)
        '''
        zongpai_list = [0, 1, 2]
        self.data = DMG_dict
        if zongpai not in zongpai_list:
            print("---- zongpai error ----")
        else:
            self.zp = zongpai

        ##根据key拿门派
        school = set()
        for key in self.data:
            school.add(int(key[0]))
        if len(school) == 1:
             self.school = school.pop()
        else:
             print("---- one time one school ,please ----")

        ##新建txt，读取cols
        folder = os.path.exists("config.txt")
        if not folder:
            f = open("config.txt", 'w')
            f.write("zp0:1;zp1:1;zp2:1")
            f.close()
            print("""---  new config has created   ---""")
        else:
            print("---  There is config!  ---")



    def write_2_excel(self):
        # 写按照宗派记录----done
        # 索引key 续写（续写不修改config,但是读取config最大值，到达config）--
        ##保存一个技能数量，用于整理
        ##自动检测文件是否存在---done
        ##否，则创建---done
        folder = os.path.exists("shuzhi_new.xlsx")
        if not folder:
            wb = Workbook()
            for i in range(2):
                wb.create_sheet(str(i))##TODO 优化sheet名称
            wb.save("shuzhi_new.xlsx")
            print("""---  new Excel "shuzhi_new.xlsx" has created   ---""")
        else:
            print("---  There is  Excel!  ---")

        wb_write = openpyxl.load_workbook("shuzhi_new.xlsx")
        sheet_names = wb_write.get_sheet_names()
        worksheet = wb_write[sheet_names[self.zp]]

        ##读config参数
        cols_max = int(read_zp(self.zp))
        #print(cols_max)
        #row_max = worksheet.max_row
        #print(row_max)

        # 走每一个cols
        #每个门派默认11行 1-11 12-22 23-33 34-44 45-55 56-66 67-77 78-88
        cols = 1

        row_num_min = (self.school-1) * 11 + 1
        #row_num_max = (self.school) * 11 + 1

        for write_cols in range(1, cols_max+1, 2):
            print(write_cols)
            if not worksheet.cell(row=row_num_min, column=write_cols).value:
                cols = write_cols
                break
            else:
                continue
        print(cols)
        log("nihao %s " % self.school)


        rows = row_num_min #根据，门派派判断# TODO 动态判断门派行数

        for key in self.data:
            col_DMG = cols + 1
            worksheet.cell(row=rows, column=cols).value = key
            worksheet.cell(row=rows, column=col_DMG).value = str(self.data[key])

            rows += 1

        wb_write.save("shuzhi_new.xlsx")
        print("----- write  done -----")
        if cols == cols_max:
            write_zp(self.zp, str(cols + 2))
        else:
            print("----- don’t need update config -----")


##分门派读
def read_from_excel(zongpai, school, cols_plus=0):
    folder = os.path.exists("shuzhi_new.xlsx")
    if not folder:
        print("""---  There is no Excel "shuzhi_new.xlsx"   ---""")
    else:
        wb_read = openpyxl.load_workbook("shuzhi_new.xlsx")
        sheet_names = wb_read.get_sheet_names()
        worksheet = wb_read[sheet_names[zongpai]]
        #row_max = worksheet.max_row
        row_num_min = (school - 1) * 11 + 1
        row_num_max = school * 11 + 1
        print(row_num_min, row_num_max)
        cols_p = cols_plus * 2
        cols = 1 + cols_p
        shuzhi_std = dict()
        for i in range(row_num_min, row_num_max):
            try:
                key = str(worksheet.cell(row=i, column=cols).value)
                DMG = worksheet.cell(row=i, column=cols+1).value
                list_num = ast.literal_eval(DMG)
                shuzhi_std[key] = list_num
            except:
                print("--- all data is here ---")
                break
    if shuzhi_std:
        print(shuzhi_std)
        return shuzhi_std

def read_from_excel_all(zongpai, cols_plus=0):
    folder = os.path.exists("shuzhi_new.xlsx")
    if not folder:
        print("""---  There is no Excel "shuzhi_new.xlsx"   ---""")
    else:
        wb_read = openpyxl.load_workbook("shuzhi_new.xlsx")
        sheet_names = wb_read.get_sheet_names()
        worksheet = wb_read[sheet_names[zongpai]]
        row_max = worksheet.max_row
        cols_p = cols_plus * 2
        cols = 1 + cols_p
        shuzhi_std_all = dict()
        for i in range(1, row_max):
            if not worksheet.cell(row=i, column=cols).value:
                continue
            else:
                key = str(worksheet.cell(row=i, column=cols).value)
                DMG = worksheet.cell(row=i, column=cols+1).value
                list_num = ast.literal_eval(DMG)
                shuzhi_std_all[key] = list_num
    if shuzhi_std_all:
        print(shuzhi_std_all)
        return shuzhi_std_all


def read_zp(zongpai):
    with open("config.txt") as f:
        con = f.readline()
        con_read = con.split(";")
        num = con_read[zongpai][4:]

    return num


def write_zp(zongpai, num):
    with open("config.txt", "r+") as f:
        con = f.readline()
        con_read = con.split(";")
        #print(con_read)

        con_read[zongpai] = "zp"+str(zongpai)+":" + str(num)
        #print(con_read)
        strall = con_read[0] +";"+con_read[1] +";"+ con_read[2]
        f.seek(0)
        f.truncate()
        f.write(strall)

    print("---- config update finish ----")



def log(*err_data):
    folder = os.path.exists("log.txt")
    if not folder:
        f = open("log.txt", 'a+')
        c_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        f.write("【"+c_time+"】" + "create log.txt"+"\n")
        f.close()
    else:
        with open("log.txt", 'a+')as f:
            c_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
            f.write("【" + c_time + "】" + str(err_data)[1:-1] + "\n")
            print("【" + c_time + "】" + str(err_data)[1:-1] + "\n")





if __name__ == '__main__':
    data1 = {
        '1_1': ['176', '92', '92', '176'],
        '1_2': ['81', '227', '228'],
        '1_3': ['138', '138', '138', '138'],
        '1_4': ['168', '216'],
        '1_5': ['109', '109', '109', '109', '109', '311'],
        '1_7': ['217', '218', '218'],
        '1_8': ['521'],
        '1_9': ['337'],
        '1_R': ['363', '363', '1064', '1064', '361', '361', '361', '361', '361', '361', '361', '361', '362', '362', '362', '362', '362', '362', '363', '363', '363', '363', '363', '363', '364', '364'],
        "1_S": ['196', '358', '358', '358', '358', '358'],
        '2_1': ['160', '160', '160'],
        '2_2': ['270'],
        '2_3': ['190'],
        '2_4': ['640'],
        '2_5': ['140', '140', '140', '140', '140', '240'],
        '2_8': ['80'],
        '2_9': ['2240'],
        '2_R': ['320', '420', '420', '320', '420', '840', '420', '1260', '420', '1680', '420', '2100', '420', '2520', '1650', '4170'],
        '2_S': ['240', '240', '240'],
        '4_b': ['580', '580', '580', '760'],
        '4_1': ['160', '160', '160'],
        '4_2': ['140', '140', '140', '140', '140'],
        '4_3': ['60', '60', '60', '60', '60', '60', '60', '60', '60', '60'],
        '4_4': ['240', '240'],
        '4_5': ['210', '210', '210', '760'],
        '4_7': ['340'],
        '4_8': ['460'],
        '4_9': ['580'],
        '4_R': ['760', '860', '860', '2140', '1280', '4280', '2140'],
        '4_S': ['240']
    }
    data2 = {
        '1_1': ['176', '92', '92', '176'],
        '1_2': ['81', '227', '228'],
        '1_3': ['138', '138', '138', '138'],
        '1_4': ['168', '216'],
        '1_5': ['109', '109', '109', '109', '109', '311'],
        '1_7': ['217', '218', '218'],
        '1_8': ['521'],
        '1_9': ['337'],
        '1_R': ['363', '363', '1064', '1064', '361', '361', '361', '361', '361', '361', '361', '361', '362', '362',
                '362', '362', '362', '362', '363', '363', '363', '363', '363', '363', '364', '364'],
        "1_S": ['196', '358', '358', '358', '358', '358'],

    }
    data3 = {
        '2_1': ['160', '160', '160'],
        '2_2': ['270'],
        '2_3': ['190'],
        '2_4': ['640'],
        '2_5': ['140', '140', '140', '140', '140', '240'],
        '2_8': ['80'],
        '2_9': ['2240'],
        '2_R': ['320', '420', '420', '320', '420', '840', '420', '1260', '420', '1680', '420', '2100', '420', '2520',
                '1650', '4170'],
        '2_S': ['240', '240', '240'],
    }
    data4 = {
        '4_b': ['580', '580', '580', '760'],
        '4_1': ['160', '160', '160'],
        '4_2': ['140', '140', '140', '140', '140'],
        '4_3': ['60', '60', '60', '60', '60', '60', '60', '60', '60', '60'],
        '4_4': ['240', '240'],
        '4_5': ['210', '210', '210', '760'],
        '4_7': ['340'],
        '4_8': ['460'],
        '4_9': ['580'],
        '4_R': ['760', '860', '860', '2140', '1280', '4280', '2140'],
        '4_S': ['240']
    }
    test = SZ_W(data3, 2)
    #
    test.write_2_excel()
    # read_from_excel(2, 2)
    # read_from_excel_all(2)
    # print(test.zp)
    # print(test.school)






